from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from database.models import Aporte, Gasto, Investimento, Projeto, Receita, TempoProjeto

# ── Estilos ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="FF1A1A1A", end_color="FF1A1A1A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11, name="Calibri")
HEADER_ALIGNMENT = Alignment(vertical="middle", horizontal="left")
BORDER_THIN = Border(bottom=Side(style="thin", color="FF444444"))
BORDER_LIGHT = Border(bottom=Side(style="thin", color="FFE0E0E0"))
ZEBRA_FILL = PatternFill(start_color="FFF5F5F5", end_color="FFF5F5F5", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, name="Calibri", color="FF1A1A1A")
SUBTITLE_FONT = Font(bold=True, size=11, name="Calibri", color="FF1A1A1A")
TOTAL_FONT = Font(bold=True, size=11, name="Calibri", color="FF1A1A1A")
TOTAL_FILL = PatternFill(start_color="FFE8E8E8", end_color="FFE8E8E8", fill_type="solid")
PROFIT_FILL = PatternFill(start_color="FFD4EDDA", end_color="FFD4EDDA", fill_type="solid")
LOSS_FILL = PatternFill(start_color="FFF8D7DA", end_color="FFF8D7DA", fill_type="solid")
MONEY_FORMAT = '#,##0.00'


def _criar_pasta(destino: Path) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)


def _estilizar_cabecalho(ws: Workbook, num_colunas: int) -> None:
    header = ws.row_dimensions[1]
    header.height = 28
    for col in range(1, num_colunas + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = BORDER_THIN


def _estilizar_linhas(ws: Workbook, inicio: int, fim: int, num_colunas: int) -> None:
    for row_idx in range(inicio, fim + 1):
        for col in range(1, num_colunas + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = Font(size=10, name="Calibri", color="FF333333")
            cell.alignment = Alignment(vertical="middle")
            cell.border = BORDER_LIGHT
            if (row_idx - inicio) % 2 == 1:
                cell.fill = ZEBRA_FILL


def _auto_largura(ws: Workbook, num_colunas: int, max_largura: int = 50) -> None:
    for col in range(1, num_colunas + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 3, 12), max_largura)


def _escrever_tabela(
    ws: Workbook,
    cabecalhos: list[str],
    linhas: list[list[object]],
    titulo: str | None = None,
) -> None:
    linha_atual = 1
    if titulo:
        ws.cell(row=1, column=1, value=titulo).font = TITLE_FONT
        linha_atual = 3

    for col_idx, cab in enumerate(cabecalhos, 1):
        ws.cell(row=linha_atual, column=col_idx, value=cab)
    _estilizar_cabecalho(ws, len(cabecalhos))

    for row_idx, row in enumerate(linhas, linha_atual + 1):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if isinstance(val, float):
                cell.number_format = MONEY_FORMAT

    _estilizar_linhas(ws, linha_atual + 1, linha_atual + len(linhas), len(cabecalhos))
    _auto_largura(ws, len(cabecalhos))


def _data_str(d: date | datetime | str | None) -> str:
    if d is None:
        return ""
    if isinstance(d, str):
        return d
    if isinstance(d, datetime):
        return d.strftime("%d/%m/%Y %H:%M")
    return d.strftime("%d/%m/%Y")


def _mes_ano(d: date | datetime | str | None) -> str:
    if d is None:
        return "Sem data"
    if isinstance(d, str):
        return d[:7]
    return d.strftime("%m/%Y")


def _add_sheet_sumario(
    wb: Workbook,
    titulo: str,
    totais: dict[str, float],
    cor: str = "FF1A1A1A",
) -> None:
    ws = wb.create_sheet(title=titulo)
    ws.cell(row=1, column=1, value=titulo).font = TITLE_FONT
    ws.cell(row=3, column=1, value="Categoria").font = SUBTITLE_FONT
    ws.cell(row=3, column=2, value="Total").font = SUBTITLE_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    ws.cell(row=3, column=1).font = HEADER_FONT
    ws.cell(row=3, column=2).fill = HEADER_FILL
    ws.cell(row=3, column=2).font = HEADER_FONT

    row = 4
    for cat, total in sorted(totais.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=cat).font = Font(size=10, name="Calibri")
        cell = ws.cell(row=row, column=2, value=round(total, 2))
        cell.number_format = MONEY_FORMAT
        cell.font = Font(size=10, name="Calibri")
        row += 1

    ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = TOTAL_FILL
    total_cell = ws.cell(row=row, column=2, value=round(sum(totais.values()), 2))
    total_cell.number_format = MONEY_FORMAT
    total_cell.font = TOTAL_FONT
    total_cell.fill = TOTAL_FILL
    _auto_largura(ws, 2)


# ── Exportações Individuais ──────────────────────────────────────────────


def exportar_projetos_para_excel(projetos: list[Projeto], destino: Path) -> Path:
    _criar_pasta(destino)
    wb = Workbook()
    ws = wb.active
    ws.title = "Projetos"

    cabecalhos = ["Cliente", "Projeto", "Tipo", "Status", "Valor total", "Valor pago", "Em aberto", "Prazo"]
    linhas: list[list[object]] = []
    for item in projetos:
        cliente_nome = item.cliente.nome if item.cliente else "Sem cliente"
        valor_total = float(item.valor_total or 0)
        valor_pago = float(item.valor_pago or 0)
        em_aberto = max(valor_total - valor_pago, 0)
        linhas.append([
            cliente_nome,
            item.titulo,
            item.tipo or "",
            item.status or "pendente",
            valor_total,
            valor_pago,
            em_aberto,
            _data_str(item.prazo),
        ])
    _escrever_tabela(ws, cabecalhos, linhas)
    wb.save(destino)
    return destino


def exportar_gastos_para_excel(gastos: list[Gasto], destino: Path) -> Path:
    _criar_pasta(destino)
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    cabecalhos = ["Data", "Descrição", "Categoria", "Valor", "Método", "Pago"]
    linhas: list[list[object]] = []
    total_por_categoria: dict[str, float] = defaultdict(float)
    total_por_mes: dict[str, float] = defaultdict(float)
    for item in gastos:
        cat = item.categoria_nome or "Sem categoria"
        valor = float(item.valor or 0)
        mes = _mes_ano(item.data)
        total_por_categoria[cat] += valor
        total_por_mes[mes] += valor
        linhas.append([
            _data_str(item.data),
            item.descricao,
            cat,
            valor,
            item.metodo or "",
            "Sim" if item.pago else "Não",
        ])
    _escrever_tabela(ws, cabecalhos, linhas)

    _add_sheet_sumario(wb, "Resumo por Categoria", dict(total_por_categoria))
    _add_sheet_sumario(wb, "Resumo por Mês", dict(total_por_mes))
    wb.save(destino)
    return destino


def exportar_receitas_para_excel(receitas: list[Receita], destino: Path) -> Path:
    _criar_pasta(destino)
    wb = Workbook()
    ws = wb.active
    ws.title = "Receitas"

    cabecalhos = ["Data", "Descrição", "Valor", "Origem", "Projeto"]
    linhas: list[list[object]] = []
    total_por_mes: dict[str, float] = defaultdict(float)
    for item in receitas:
        valor = float(item.valor or 0)
        mes = _mes_ano(item.data)
        total_por_mes[mes] += valor
        linhas.append([
            _data_str(item.data),
            item.descricao,
            valor,
            item.origem or "",
            item.projeto.titulo if item.projeto else "",
        ])
    _escrever_tabela(ws, cabecalhos, linhas)
    _add_sheet_sumario(wb, "Resumo por Mês", dict(total_por_mes))
    wb.save(destino)
    return destino


def exportar_investimentos_para_excel(
    investimentos: list[Investimento],
    aportes: list[Aporte],
    destino: Path,
) -> Path:
    _criar_pasta(destino)
    wb = Workbook()
    ws = wb.active
    ws.title = "Investimentos"

    def total_investido(inv_id: int) -> float:
        total = 0.0
        for ap in aportes:
            if ap.investimento_id != inv_id:
                continue
            if ap.tipo == "resgate":
                total -= float(ap.valor or 0)
            else:
                total += float(ap.valor or 0)
        return total

    cabecalhos = ["Nome", "Tipo", "Ativo", "Total investido", "Meta valor", "Meta data"]
    linhas: list[list[object]] = []
    for item in investimentos:
        linhas.append([
            item.nome,
            item.tipo,
            "Sim" if item.ativo else "Não",
            round(total_investido(item.id), 2),
            float(item.meta_valor or 0),
            _data_str(item.meta_data),
        ])
    _escrever_tabela(ws, cabecalhos, linhas)
    wb.save(destino)
    return destino


def exportar_aportes_para_excel(aportes: list[Aporte], destino: Path) -> Path:
    _criar_pasta(destino)
    wb = Workbook()
    ws = wb.active
    ws.title = "Aportes"

    cabecalhos = ["Data", "Investimento", "Tipo", "Valor", "Notas"]
    linhas: list[list[object]] = []
    total_por_mes: dict[str, float] = defaultdict(float)
    for item in aportes:
        valor = float(item.valor or 0)
        mes = _mes_ano(item.data)
        total_por_mes[mes] += valor
        linhas.append([
            _data_str(item.data),
            item.investimento.nome if item.investimento else "",
            item.tipo or "aporte",
            valor,
            item.notas or "",
        ])
    _escrever_tabela(ws, cabecalhos, linhas)
    _add_sheet_sumario(wb, "Resumo por Mês", dict(total_por_mes))
    wb.save(destino)
    return destino


def exportar_sessoes_para_excel(sessoes: list[TempoProjeto], destino: Path) -> Path:
    _criar_pasta(destino)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sessões"

    cabecalhos = ["Projeto", "Início", "Fim", "Duração (min)", "Descrição"]
    linhas: list[list[object]] = []
    for sessao in sessoes:
        fim = sessao.fim
        duracao = sessao.duracao_min
        if duracao is None and fim is not None:
            duracao = max(int((fim - sessao.inicio).total_seconds() // 60), 0)
        linhas.append([
            sessao.projeto.titulo if sessao.projeto else "-",
            _data_str(sessao.inicio),
            _data_str(fim),
            duracao or 0,
            sessao.descricao or "",
        ])
    _escrever_tabela(ws, cabecalhos, linhas)
    wb.save(destino)
    return destino


# ── Relatório Consolidado ────────────────────────────────────────────────


def exportar_relatorio_completo(
    projetos: list[Projeto],
    gastos: list[Gasto],
    receitas: list[Receita],
    investimentos: list[Investimento],
    aportes: list[Aporte],
    sessoes: list[TempoProjeto],
    destino: Path,
) -> Path:
    _criar_pasta(destino)
    wb = Workbook()

    # Sheet: Lucro/Prejuízo
    ws = wb.active
    ws.title = "Lucro e Prejuízo"

    total_receitas = sum(float(r.valor or 0) for r in receitas)
    total_gastos = sum(float(g.valor or 0) for g in gastos)
    saldo = total_receitas - total_gastos

    ws.cell(row=1, column=1, value="Relatório Consolidado - Lucro e Prejuízo").font = TITLE_FONT
    ws.cell(row=3, column=1, value="Indicador").font = SUBTITLE_FONT
    ws.cell(row=3, column=2, value="Valor").font = SUBTITLE_FONT
    for c in [1, 2]:
        ws.cell(row=3, column=c).fill = HEADER_FILL
        ws.cell(row=3, column=c).font = HEADER_FONT

    data = [
        ("Total de Receitas", round(total_receitas, 2)),
        ("Total de Gastos", round(total_gastos, 2)),
        ("Saldo (Receitas - Gastos)", round(saldo, 2)),
        ("Total em Projetos", round(sum(float(p.valor_total or 0) for p in projetos), 2)),
        ("Total Recebido de Projetos", round(sum(float(p.valor_pago or 0) for p in projetos), 2)),
        ("Saldo em Aberto (Projetos)", round(sum(max(float(p.valor_total or 0) - float(p.valor_pago or 0), 0) for p in projetos), 2)),
        (
            "Total Horas Registradas",
            round(sum(s.duracao_min or 0 for s in sessoes) / 60, 1),
        ),
    ]
    for i, (label, val) in enumerate(data, 4):
        ws.cell(row=i, column=1, value=label).font = Font(size=10, name="Calibri")
        cell = ws.cell(row=i, column=2, value=val)
        if isinstance(val, float):
            cell.number_format = MONEY_FORMAT
        cell.font = Font(size=10, name="Calibri")

    saldo_row = 6
    fill = PROFIT_FILL if saldo >= 0 else LOSS_FILL
    ws.cell(row=saldo_row, column=1).fill = fill
    ws.cell(row=saldo_row, column=2).fill = fill
    _auto_largura(ws, 2, 60)

    # Sheet: Receitas por mês
    rec_mes: dict[str, float] = defaultdict(float)
    for r in receitas:
        rec_mes[_mes_ano(r.data)] += float(r.valor or 0)
    _add_sheet_sumario(wb, "Receitas por Mês", dict(rec_mes))

    # Sheet: Gastos por categoria
    gasto_cat: dict[str, float] = defaultdict(float)
    for g in gastos:
        gasto_cat[g.categoria_nome or "Sem categoria"] += float(g.valor or 0)
    _add_sheet_sumario(wb, "Gastos por Categoria", dict(gasto_cat))

    # Sheet: Gastos por mês
    gasto_mes: dict[str, float] = defaultdict(float)
    for g in gastos:
        gasto_mes[_mes_ano(g.data)] += float(g.valor or 0)
    _add_sheet_sumario(wb, "Gastos por Mês", dict(gasto_mes))

    # Sheet: Projetos
    ws_proj = wb.create_sheet(title="Projetos")
    cab_proj = ["Cliente", "Projeto", "Status", "Valor total", "Valor pago", "Em aberto", "Prazo"]
    linhas_proj: list[list[object]] = []
    for p in projetos:
        em_aberto = max(float(p.valor_total or 0) - float(p.valor_pago or 0), 0)
        linhas_proj.append([
            p.cliente.nome if p.cliente else "Sem cliente",
            p.titulo,
            p.status or "pendente",
            float(p.valor_total or 0),
            float(p.valor_pago or 0),
            em_aberto,
            _data_str(p.prazo),
        ])
    _escrever_tabela(ws_proj, cab_proj, linhas_proj)

    # Sheet: Gastos detalhado
    ws_gastos = wb.create_sheet(title="Gastos")
    cab_gastos = ["Data", "Descrição", "Categoria", "Valor", "Método", "Pago"]
    linhas_gastos: list[list[object]] = []
    for g in gastos:
        linhas_gastos.append([
            _data_str(g.data), g.descricao, g.categoria_nome or "Sem categoria",
            float(g.valor or 0), g.metodo or "", "Sim" if g.pago else "Não",
        ])
    _escrever_tabela(ws_gastos, cab_gastos, linhas_gastos)

    # Sheet: Receitas detalhado
    ws_rec = wb.create_sheet(title="Receitas")
    cab_rec = ["Data", "Descrição", "Valor", "Origem", "Projeto"]
    linhas_rec: list[list[object]] = []
    for r in receitas:
        linhas_rec.append([
            _data_str(r.data), r.descricao, float(r.valor or 0),
            r.origem or "", r.projeto.titulo if r.projeto else "",
        ])
    _escrever_tabela(ws_rec, cab_rec, linhas_rec)

    # Sheet: Sessões
    ws_sess = wb.create_sheet(title="Sessões")
    cab_sess = ["Projeto", "Início", "Fim", "Duração (min)", "Descrição"]
    linhas_sess: list[list[object]] = []
    for s in sessoes:
        duracao = s.duracao_min
        if duracao is None and s.fim is not None:
            duracao = max(int((s.fim - s.inicio).total_seconds() // 60), 0)
        linhas_sess.append([
            s.projeto.titulo if s.projeto else "-",
            _data_str(s.inicio), _data_str(s.fim),
            duracao or 0, s.descricao or "",
        ])
    _escrever_tabela(ws_sess, cab_sess, linhas_sess)

    wb.save(destino)
    return destino


# ── PDF ──────────────────────────────────────────────────────────────────


def _add_pdf_table(
    pdf: Any,
    cabecalhos: list[str],
    linhas: list[list[str]],
    col_widths: list[int] | None = None,
) -> None:
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    data = [cabecalhos] + [[str(c) for c in row] for row in linhas]
    if col_widths is None:
        col_widths = [max(80, len(h) * 7) for h in cabecalhos]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1A1A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    table.setStyle(TableStyle(style))
    pdf.append(table)


def exportar_relatorio_pdf(
    projetos: list[Projeto],
    gastos: list[Gasto],
    receitas: list[Receita],
    investimentos: list[Investimento],
    aportes: list[Aporte],
    sessoes: list[TempoProjeto],
    destino: Path,
) -> Path:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    _criar_pasta(destino)
    doc = SimpleDocTemplate(
        str(destino),
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    estilos = getSampleStyleSheet()
    elementos: list[object] = []

    # Título
    elementos.append(Paragraph("Relatório Consolidado DevFlow", estilos["Title"]))
    elementos.append(Spacer(1, 6 * mm))
    elementos.append(
        Paragraph(
            f"Gerado em: {datetime.now():%d/%m/%Y %H:%M}",
            estilos["Normal"],
        )
    )
    elementos.append(Spacer(1, 10 * mm))

    # Resumo Financeiro
    elementos.append(Paragraph("Resumo Financeiro", estilos["Heading2"]))
    total_rec = sum(float(r.valor or 0) for r in receitas)
    total_gas = sum(float(g.valor or 0) for g in gastos)
    saldo = total_rec - total_gas
    total_proj = sum(float(p.valor_total or 0) for p in projetos)
    total_pago = sum(float(p.valor_pago or 0) for p in projetos)
    elementos.append(
        Table(
            [
                ["Indicador", "Valor"],
                ["Total Receitas", f"{total_rec:,.2f}"],
                ["Total Gastos", f"{total_gas:,.2f}"],
                ["Saldo", f"{saldo:,.2f}"],
                ["Total Projetos", f"{total_proj:,.2f}"],
                ["Total Recebido", f"{total_pago:,.2f}"],
            ],
            colWidths=[180, 100],
            repeatRows=1,
        )
    )
    elementos.append(Spacer(1, 8 * mm))

    # Projetos
    if projetos:
        elementos.append(Paragraph("Projetos", estilos["Heading2"]))
        linhas_proj: list[list[str]] = []
        for p in projetos:
            cliente = p.cliente.nome if p.cliente else "Sem cliente"
            em_aberto = max(float(p.valor_total or 0) - float(p.valor_pago or 0), 0)
            linhas_proj.append([
                cliente,
                p.titulo,
                p.status or "pendente",
                f"{float(p.valor_total or 0):,.2f}",
                f"{float(p.valor_pago or 0):,.2f}",
                f"{em_aberto:,.2f}",
            ])
        _add_pdf_table(
            elementos,
            ["Cliente", "Projeto", "Status", "Total", "Pago", "Aberto"],
            linhas_proj,
            col_widths=[80, 90, 50, 50, 50, 50],
        )
        elementos.append(PageBreak())

    # Gastos
    if gastos:
        elementos.append(Paragraph("Gastos", estilos["Heading2"]))
        linhas_gas: list[list[str]] = []
        for g in gastos[:50]:
            linhas_gas.append([
                _data_str(g.data),
                g.descricao[:40],
                g.categoria_nome or "-",
                f"{float(g.valor or 0):,.2f}",
                "Sim" if g.pago else "Não",
            ])
        _add_pdf_table(
            elementos,
            ["Data", "Descrição", "Categoria", "Valor", "Pago"],
            linhas_gas,
            col_widths=[50, 100, 60, 50, 40],
        )
        elementos.append(Spacer(1, 6 * mm))

    # Receitas
    if receitas:
        elementos.append(Paragraph("Receitas", estilos["Heading2"]))
        linhas_rec: list[list[str]] = []
        for r in receitas[:50]:
            linhas_rec.append([
                _data_str(r.data),
                r.descricao[:40],
                r.origem or "-",
                f"{float(r.valor or 0):,.2f}",
            ])
        _add_pdf_table(
            elementos,
            ["Data", "Descrição", "Origem", "Valor"],
            linhas_rec,
            col_widths=[50, 110, 60, 50],
        )

    doc.build(elementos)
    return destino
